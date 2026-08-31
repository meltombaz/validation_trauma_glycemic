# ============================================================
# Streamlit app — Diabetes / Prediabetes Screening
# Reduced sex-specific feature models
#
# Run with:
#   streamlit run app_diabetes_screening.py
#
# Expected model files:
#   final_reduced_model_Female.pkl
#   final_reduced_model_Male.pkl
# ============================================================

import io
import os
import joblib
import openpyxl
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components


# ============================================================
# CONFIG
# ============================================================

MODEL_DIR = os.environ.get("MODEL_DIR", ".")

MODEL_PATHS = {
    "Female": os.path.join(
        MODEL_DIR,
        "final_reduced_model_Female.pkl"
    ),
    "Male": os.path.join(
        MODEL_DIR,
        "final_reduced_model_Male.pkl"
    ),
}


FEATURE_INFO = {
    "Age": {
        "label": "Age",
        "unit": "years",
        "group": "Anthropometrics",
    },

    "LEU": {
        "label": "Leukocytes",
        "unit": "1/\u00b5L",
        "group": "Laboratory values",
    },

    "MCHC": {
        "label": "MCHC",
        "unit": "g/dL",
        "group": "Laboratory values",
    },

    "MeanPlateletVolume": {
        "label": "Mean Platelet Volume",
        "unit": "fL",
        "group": "Laboratory values",
    },

    "QUICK": {
        "label": "QUICK",
        "unit": "%",
        "group": "Laboratory values",
    },

    "WaistCircumference": {
        "label": "Waist Circumference",
        "unit": "cm",
        "group": "Anthropometrics",
    },

    "APTT": {
        "label": "APTT",
        "unit": "sec",
        "group": "Laboratory values",
    },

    "BMI": {
        "label": "BMI",
        "unit": "kg/m\u00b2",
        "computed_from_weight_height": True,
        "group": "Anthropometrics",
    },

    "Previous High Blood Sugar Levels": {
        "label": "Previous High Blood Sugar Levels",
        "categorical_labels": {
            2: "Yes",
            0: "No",
        },
        "group": "Medical history",
    },

    "High Blood Pressure Medicine": {
        "label": "High Blood Pressure Medicine",
        "categorical_labels": {
            2: "Yes",
            0: "No",
        },
        "group": "Medical history",
    },
}


st.set_page_config(
    page_title="Glycemic Risk Screening",
    page_icon="\U0001fa7a",
    layout="wide",
    initial_sidebar_state="collapsed",
)


def block_browser_forced_dark_mode():
    """Chrome/Edge/Android's built-in 'Auto Dark Mode' heuristically
    re-colors a page's rendered pixels AFTER our CSS has already been
    applied — it isn't part of the normal CSS cascade, so no amount of
    our own CSS (including `color-scheme: light`) can reliably stop it.
    The one thing the browser explicitly checks for is a
    `<meta name="color-scheme">` tag in <head>, which `st.html()` can't
    reach (it only injects into <body>). A same-origin Streamlit
    component running in an iframe *can* reach the real parent
    document, so we use one just to append that single meta tag."""
    components.html(
        """
        <script>
        try {
            var doc = window.parent.document;
            if (!doc.querySelector('meta[name="color-scheme"]')) {
                var meta = doc.createElement('meta');
                meta.name = 'color-scheme';
                meta.content = 'light only';
                doc.head.appendChild(meta);
            }
        } catch (e) {}
        </script>
        """,
        height=0,
    )


block_browser_forced_dark_mode()


# ============================================================
# STYLING
# ============================================================

def inject_css():
    st.html("""
    <style>

    /* ---------------------------------------------------------
       FORCE LIGHT MODE
       Belt-and-braces alongside .streamlit/config.toml (base="light"):
       this app's colors are hardcoded for a light background, so if a
       hosting environment ever ignores config.toml and applies a dark
       theme anyway, text/backgrounds must not be allowed to flip.
    --------------------------------------------------------- */

    :root {
        color-scheme: light only;
    }

    .stApp, .stApp * {
        color-scheme: light only;
    }

    body, .stApp {
        background: #f5f7fa !important;
        color: #172033 !important;
    }

    [data-testid="stAppViewContainer"],
    [data-testid="stHeader"],
    [data-testid="stSidebar"] {
        background: #f5f7fa !important;
        color: #172033 !important;
    }

    /* Native widget text/labels/values that would otherwise pick up
       a dark-mode text color from the browser/OS */
    label[data-testid="stWidgetLabel"] p,
    [data-testid="stMarkdownContainer"] p,
    [data-testid="stMarkdownContainer"] li,
    [data-testid="stCaptionContainer"],
    [data-testid="stMetricLabel"],
    [data-testid="stMetricValue"],
    [data-testid="stText"] {
        color: #172033 !important;
    }

    div[data-baseweb="input"] input,
    div[data-baseweb="select"] *,
    div[data-baseweb="base-input"] input,
    textarea {
        color: #172033 !important;
        background: #fafbfd !important;
    }


    /* ---------------------------------------------------------
       PAGE
    --------------------------------------------------------- */

    .stApp {
        background: #f5f7fa;
    }

    .block-container {
        max-width: 1320px;
        padding-top: 1.5rem;
        padding-left: 2.5rem;
        padding-right: 2.5rem;
        padding-bottom: 4rem;
    }

    /* Reduce some default Streamlit vertical whitespace */
    div[data-testid="stVerticalBlock"] {
        gap: 0.65rem;
    }


    /* ---------------------------------------------------------
       HEADER
    --------------------------------------------------------- */

    .app-header {
        background: #004994;
        border: 1px solid #e4e9f0;
        border-radius: 18px;
        padding: 25px 30px;
        margin-bottom: 18px;

        box-shadow:
            0 1px 2px rgba(15, 23, 42, 0.02),
            0 6px 20px rgba(15, 23, 42, 0.035);
    }

    .header-top {
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 20px;
    }

    .app-title {
        font-size: 2rem;
        font-weight: 750;
        color: #ffffff;
        letter-spacing: -0.025em;
        line-height: 1.15;
    }

    .app-subtitle {
        margin-top: 7px;
        color: #ffffff;
        font-size: 0.94rem;
    }

    .research-badge {
        flex-shrink: 0;
        background: #edf4ff;
        color: #315f9f;
        border: 1px solid #d5e4fa;
        padding: 6px 11px;
        border-radius: 999px;
        font-weight: 700;
        font-size: 0.70rem;
        letter-spacing: 0.04em;
    }

    .disclaimer {
        margin-top: 17px;
        padding-top: 14px;
        border-top: 1px solid #edf0f4;
        color: #ffffff;
        font-size: 0.82rem;
    }

    .disclaimer-dot {
        display: inline-block;
        width: 7px;
        height: 7px;
        border-radius: 50%;
        background: #7b91ad;
        margin-right: 7px;
    }


    /* ---------------------------------------------------------
       SECTION HEADINGS
    --------------------------------------------------------- */

    .step-label {
        font-size: 0.72rem;
        font-weight: 700;
        color: #6580a3;
        letter-spacing: 0.07em;
        text-transform: uppercase;
        margin-bottom: 3px;
    }

    .section-title {
        color: #1d2939;
        font-size: 1.15rem;
        font-weight: 700;
        margin-bottom: 2px;
    }

    .section-subtitle {
        color: #7a8798;
        font-size: 0.82rem;
        margin-bottom: 12px;
    }


    /* ---------------------------------------------------------
       STREAMLIT CONTAINERS AS CARDS
    --------------------------------------------------------- */

    div[data-testid="stVerticalBlockBorderWrapper"] {
        background: #ffffff;
        border-color: #e3e8ef !important;
        border-radius: 15px !important;

        box-shadow:
            0 1px 2px rgba(15, 23, 42, 0.015),
            0 4px 14px rgba(15, 23, 42, 0.025);
    }


    /* ---------------------------------------------------------
       INPUTS
    --------------------------------------------------------- */

    div[data-baseweb="input"] > div,
    div[data-baseweb="select"] > div {
        background: #fafbfd;
        border-radius: 9px !important;
    }

    label[data-testid="stWidgetLabel"] p {
        font-size: 0.82rem;
        color: #344054;
        font-weight: 500;
    }

    div[data-testid="stNumberInput"] input {
        font-size: 0.88rem;
    }


    /* ---------------------------------------------------------
       BUTTON
    --------------------------------------------------------- */

    div.stButton > button[kind="primary"] {
        height: 3.2rem;
        border-radius: 11px;
        font-size: 0.95rem;
        font-weight: 700;
        box-shadow: none;
    }


    /* ---------------------------------------------------------
       METRICS
    --------------------------------------------------------- */

    [data-testid="stMetric"] {
        background: transparent;
        border: 0;
        padding: 0;
    }

    [data-testid="stMetricLabel"] {
        font-size: 0.75rem;
    }

    [data-testid="stMetricValue"] {
        font-size: 1.35rem;
    }


    /* ---------------------------------------------------------
       RESULT
    --------------------------------------------------------- */

    .result-card {
        background: #ffffff;
        border: 1px solid #dfe5ec;
        border-radius: 18px;
        padding: 28px 30px;
        margin-top: 12px;

        box-shadow:
            0 4px 20px rgba(15, 23, 42, 0.045);
    }

    .result-label {
        color: #718096;
        font-size: 0.75rem;
        text-transform: uppercase;
        letter-spacing: 0.06em;
        font-weight: 700;
    }

    .result-probability {
        color: #172033;
        font-size: 3.2rem;
        font-weight: 760;
        letter-spacing: -0.04em;
        margin-top: 3px;
    }

    .result-classification {
        color: #344054;
        font-size: 1rem;
        margin-top: 7px;
    }

    .status-above,
    .status-below {
        display: inline-block;
        margin-top: 13px;
        border-radius: 999px;
        padding: 6px 11px;
        font-size: 0.78rem;
        font-weight: 650;
    }

    .status-above {
        background: #fff4e8;
        color: #99510d;
        border: 1px solid #f4d4af;
    }

    .status-below {
        background: #eef7f3;
        color: #287055;
        border: 1px solid #cee6da;
    }

    .result-note {
        margin-top: 15px;
        color: #7b8797;
        font-size: 0.78rem;
    }

    </style>
    """)

inject_css()


# ============================================================
# MODEL FUNCTIONS
# ============================================================

@st.cache_resource
def load_artifact(path):
    return joblib.load(path)


def field_label(feat):
    info = FEATURE_INFO.get(feat, {})
    label = info.get("label", feat)
    unit = info.get("unit")

    if unit:
        return f"{label} ({unit})"

    return label


def field_label_no_unit(feat):
    return FEATURE_INFO.get(feat, {}).get("label", feat)


def predict(artifact, X_row):

    X = X_row.copy()

    num_in = artifact["num_features"]
    cat_in = artifact["cat_features"]

    if num_in and artifact["num_imputer"] is not None:

        X[num_in] = artifact["num_imputer"].transform(
            X[num_in]
        )

        if artifact["scaler"] is not None:
            X[num_in] = artifact["scaler"].transform(
                X[num_in]
            )

    if cat_in and artifact["cat_imputer"] is not None:

        X[cat_in] = artifact["cat_imputer"].transform(
            X[cat_in]
        )

    X = X[artifact["feature_names"]]

    probability = float(
        artifact["model"]
        .predict_proba(X)[:, 1][0]
    )

    return probability


# ============================================================
# BATCH UPLOAD (Excel database)
#
# Reads the "Datentabelle" sheet of the HbA1c database file (one
# column per patient, one row per questionnaire/lab variable) and
# scores every complete patient using the exact same predict()
# pipeline as the single-patient form above.
# ============================================================

BATCH_ROW = {
    "survey_date": 3,
    "birth_date": 5,
    "age_calc": 6,               # =DATEDIF(birth, survey, "Y")
    "sex": 7,                    # 'weiblich' / 'männlich' / 'divers'
    "height_m": 11,               # in meters
    "weight_kg": 12,
    "bmi_calc": 13,               # =weight / height^2
    "waist_cm": 14,               # Taillenumfang
    "bp_medicine": 18,            # Blutdruckmedikamente jemals verordnet?
    "high_blood_sugar": 19,       # Jemals zu hohe Blutzuckerwerte festgestellt?
    "leukocytes": 83,             # Leukozyten, 1/µL
    "mchc": 88,                   # g/dL
    "quick": 94,                  # %
    "aptt": 96,                   # sec
    "mean_platelet_volume": 93,   # mittleres Plättchenvolumen, fL
}
BATCH_STUDENT_ID_ROW = 2
BATCH_FIRST_PATIENT_COL = 8   # column H — column G holds the 'Hexxx' template label
BATCH_PLACEHOLDER_VALUES = {None, "", "bitte auswählen", "Hexxx", "keine Angabe notwendig"}
BATCH_YES_NO_MAP = {"ja": 2, "nein": 0}
BATCH_SEX_MODEL_MAP = {"weiblich": "Female", "männlich": "Male"}


def batch_is_missing(value):
    if value is None:
        return True
    if isinstance(value, str) and value.strip() in BATCH_PLACEHOLDER_VALUES:
        return True
    return False


def batch_find_patient_columns(ws):
    cols = []
    for c in range(BATCH_FIRST_PATIENT_COL, ws.max_column + 1):
        sid = ws.cell(row=BATCH_STUDENT_ID_ROW, column=c).value
        if sid and str(sid).strip() not in BATCH_PLACEHOLDER_VALUES:
            cols.append((c, str(sid).strip()))
    return cols


def batch_extract_patient(ws, col):
    """Pull one patient's column out of the Datentabelle sheet and
    translate it into the model's feature names. Returns
    (features_dict, sex_label, problems)."""

    problems = []

    def cell(row_key):
        return ws.cell(row=BATCH_ROW[row_key], column=col).value

    sex_raw = cell("sex")
    if batch_is_missing(sex_raw):
        problems.append("Sex ('Geschlecht') is missing.")
        sex_label = None
    else:
        sex_label = BATCH_SEX_MODEL_MAP.get(str(sex_raw).strip().lower())
        if sex_label is None:
            problems.append(
                f"Sex value '{sex_raw}' is not 'weiblich'/'männlich' — "
                "no matching sex-specific model."
            )

    age = cell("age_calc")
    if batch_is_missing(age) or not isinstance(age, (int, float)):
        birth = cell("birth_date")
        survey = cell("survey_date")
        if hasattr(birth, "year") and hasattr(survey, "year"):
            age = survey.year - birth.year - (
                (survey.month, survey.day) < (birth.month, birth.day)
            )
        else:
            problems.append("Age could not be determined (missing birth/survey date).")
            age = None

    bmi = cell("bmi_calc")
    if batch_is_missing(bmi) or isinstance(bmi, str):
        weight = cell("weight_kg")
        height_m = cell("height_m")
        if isinstance(weight, (int, float)) and isinstance(height_m, (int, float)) and height_m > 0:
            bmi = weight / (height_m ** 2)
        else:
            problems.append("BMI could not be computed (missing weight/height).")
            bmi = None

    numeric_fields = {
        "WaistCircumference": "waist_cm",
        "LEU": "leukocytes",
        "MCHC": "mchc",
        "QUICK": "quick",
        "APTT": "aptt",
        "MeanPlateletVolume": "mean_platelet_volume",
    }

    features = {"Age": age, "BMI": bmi}

    for feat_name, row_key in numeric_fields.items():
        value = cell(row_key)
        if batch_is_missing(value) or not isinstance(value, (int, float)):
            problems.append(f"{field_label_no_unit(feat_name)} is missing or non-numeric.")
            features[feat_name] = None
        else:
            features[feat_name] = float(value)

    yes_no_fields = {
        "Previous High Blood Sugar Levels": "high_blood_sugar",
        "High Blood Pressure Medicine": "bp_medicine",
    }

    for feat_name, row_key in yes_no_fields.items():
        value = cell(row_key)
        if batch_is_missing(value):
            problems.append(f"{field_label_no_unit(feat_name)} is missing.")
            features[feat_name] = None
        else:
            key = str(value).strip().lower()
            if key not in BATCH_YES_NO_MAP:
                problems.append(f"{field_label_no_unit(feat_name)} value '{value}' is not 'ja'/'nein'.")
                features[feat_name] = None
            else:
                features[feat_name] = BATCH_YES_NO_MAP[key]

    return features, sex_label, problems


def run_batch(workbook_bytes, threshold):
    """Score every patient column found in the uploaded workbook.
    Returns a results DataFrame (one row per patient column)."""

    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True)

    if "Datentabelle" not in wb.sheetnames:
        return None, "No 'Datentabelle' sheet found in the uploaded file."

    ws = wb["Datentabelle"]
    patient_cols = batch_find_patient_columns(ws)

    if not patient_cols:
        return pd.DataFrame(), None

    artifacts = {}
    results = []

    for col, student_id in patient_cols:
        features, sex_label, problems = batch_extract_patient(ws, col)

        row = {
            "Student-ID": student_id,
            "Sex": sex_label or "?",
            "Status": "OK" if not problems and sex_label else "Incomplete",
            "Probability": None,
            "Classification": None,
            "Issues": "; ".join(problems),
        }

        if row["Status"] == "OK":
            model_path = MODEL_PATHS[sex_label]
            if not os.path.exists(model_path):
                row["Status"] = "Incomplete"
                row["Issues"] = f"Model file not found: {model_path}"
            else:
                if sex_label not in artifacts:
                    artifacts[sex_label] = load_artifact(model_path)

                artifact = artifacts[sex_label]
                X_row = pd.DataFrame([features])[artifact["feature_names"]]
                probability = predict(artifact, X_row)
                row["Probability"] = round(probability, 4)
                row["Classification"] = (
                    "Pre-diabetes / Diabetes risk"
                    if probability >= threshold else "Normal"
                )

        results.append(row)

    return pd.DataFrame(results), None


def to_excel_bytes(df):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Predictions", index=False)
    return buffer.getvalue()


def render_batch_tab():

    st.markdown(
        '<div class="section-heading">Upload patient database</div>',
        unsafe_allow_html=True,
    )

    st.caption(
        "Upload the HbA1c database Excel file (\u201cDatentabelle\u201d sheet). "
        "Each patient column is scored with the same sex-specific model "
        "used in the single-patient form."
    )

    threshold = st.slider(
        "Decision threshold",
        min_value=0.0,
        max_value=1.0,
        value=0.50,
        step=0.01,
        key="batch_threshold",
        help=(
            "Probability threshold used to convert each patient's "
            "model probability into a binary classification."
        ),
    )

    uploaded = st.file_uploader(
        "Excel database (.xlsx)",
        type=["xlsx"],
        key="batch_uploader",
    )

    if uploaded is None:
        return

    results_df, error = run_batch(uploaded.getvalue(), threshold)

    if error:
        st.error(error)
        return

    if results_df.empty:
        st.warning("No patient columns were found in the uploaded file.")
        return

    ok_count = int((results_df["Status"] == "OK").sum())
    total = len(results_df)

    st.write("")
    c1, c2 = st.columns(2)
    with c1:
        st.metric("Patients scored", f"{ok_count} / {total}")
    with c2:
        if ok_count:
            flagged = int(
                (results_df["Classification"] == "Pre-diabetes / Diabetes risk").sum()
            )
            st.metric("Above threshold", flagged)

    st.write("")
    st.dataframe(
        results_df,
        hide_index=True,
        use_container_width=True,
    )

    incomplete = total - ok_count
    if incomplete:
        st.caption(
            f"{incomplete} patient column(s) were skipped — see the "
            "\u201cIssues\u201d column for what's missing or unreadable."
        )

    st.download_button(
        "Download results (.xlsx)",
        data=to_excel_bytes(results_df),
        file_name="predictions.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )


# ============================================================
# VALIDATION HELPERS
#
# Plain st.number_input silently clamps out-of-range typed values to
# min_value/max_value instead of telling the user why (e.g. a height
# of 1.70 would quietly become 100.0). To catch and explain entry
# errors instead, numeric fields below are collected as free text and
# validated explicitly: this also lets us accept both English (.) and
# German (,) decimal separators, and flag likely unit-confusion (e.g.
# height typed in meters instead of centimeters) with a clear message
# rather than a silent, wrong number.
# ============================================================

def parse_locale_number(raw):
    """Parse a user-typed number, accepting both '.' and ',' as the
    decimal separator. Returns (value, error_message); error_message
    is None on success."""
    if raw is None:
        return None, "This field is required."

    s = str(raw).strip()
    if s == "":
        return None, "This field is required."

    s = s.replace(" ", "")
    has_comma = "," in s
    has_dot = "." in s

    if has_comma and has_dot:
        # Two separators present: the LAST one is the decimal
        # separator, the earlier one is a thousands separator.
        # "1.234,56" (German) -> 1234.56 | "1,234.56" (English) -> 1234.56
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif has_comma:
        # Only a comma present -> treat as a decimal separator (German style)
        s = s.replace(",", ".")
    # else: only a dot, or no separator -> already a valid float string

    try:
        value = float(s)
    except ValueError:
        return None, (
            f'Could not read "{raw}" as a number. '
            f"Use digits and a decimal point or comma, e.g. 72.5 or 72,5."
        )

    return value, None


def validate_positive(value, label, unit=""):
    """Reject zero/negative values, which are never physiologically
    valid for the fields this app collects."""
    if value <= 0:
        unit_str = f" {unit}" if unit else ""
        return f"{label} must be greater than 0{unit_str}."
    return None


def validate_height_cm(value):
    """Flag values that look like they were entered in meters
    (e.g. 1.70) rather than centimeters (170)."""
    err = validate_positive(value, "Height", "cm")
    if err:
        return err
    if value < 10:
        return (
            f"This looks like it may be in meters ({value:g}). "
            f"Please enter height in centimeters, e.g. 170."
        )
    return None


# ============================================================
# INPUT COMPONENTS
# ============================================================

def numeric_input(feat, stats, sex, errors):

    minimum = float(stats["min"])
    maximum = float(stats["max"])
    median = float(stats["median"])
    label = field_label(feat)

    raw = st.text_input(
        label,
        value=f"{median:.2f}",
        help=(
            f"Observed training range: "
            f"{minimum:.2f} to {maximum:.2f}"
        ),
        key=f"{sex}_{feat}",
    )

    value, err = parse_locale_number(raw)
    if err is None:
        err = validate_positive(value, field_label_no_unit(feat))

    if err:
        st.error(err)
        errors.append(f"{label}: {err}")
        return None

    if value < minimum or value > maximum * 1.5:
        st.warning(
            f"{value:g} is outside the observed training range "
            f"({minimum:.2f} to {maximum:.2f}). Double-check this value."
        )

    return value


def categorical_input(feat, stats, sex, errors):

    info = FEATURE_INFO.get(feat, {})

    raw_options = stats["values"]

    labels = info.get(
        "categorical_labels",
        {}
    )

    display_options = [
        str(labels.get(value, value))
        for value in raw_options
    ]

    choice = st.selectbox(
        field_label(feat),
        display_options,
        key=f"{sex}_{feat}",
    )

    return raw_options[
        display_options.index(choice)
    ]


def bmi_input(feat, stats, sex, errors):

    st.html(
        """
        <div style="
            font-size:0.88rem;
            font-weight:600;
            color:#334155;
            margin-bottom:8px;
        ">
        BMI
        </div>
        """
    )

    c1, c2 = st.columns(2)

    with c1:
        weight_raw = st.text_input(
            "Weight (kg)",
            value="75.0",
            key=f"{sex}_weight",
        )
        weight_kg, w_err = parse_locale_number(weight_raw)
        if w_err is None:
            w_err = validate_positive(weight_kg, "Weight", "kg")
        if w_err:
            st.error(w_err)
            errors.append(f"Weight: {w_err}")

    with c2:
        height_raw = st.text_input(
            "Height (cm)",
            value="170.0",
            key=f"{sex}_height",
        )
        height_cm, h_err = parse_locale_number(height_raw)
        if h_err is None:
            h_err = validate_height_cm(height_cm)
        if h_err:
            st.error(h_err)
            errors.append(f"Height: {h_err}")

    if w_err or h_err:
        st.caption("Fix the value(s) above to calculate BMI.")
        return None

    bmi = weight_kg / (
        (height_cm / 100) ** 2
    )

    c1, c2 = st.columns([1, 1.5])

    with c1:
        st.metric(
            "Calculated BMI",
            f"{bmi:.1f} kg/m\u00b2",
        )

    with c2:
        st.caption(
            "Training range: "
            f"{stats['min']:.1f} to "
            f"{stats['max']:.1f} kg/m\u00b2"
        )
        if bmi < stats["min"] or bmi > stats["max"] * 1.5:
            st.warning(
                f"Calculated BMI ({bmi:.1f}) is outside the observed "
                f"training range. Double-check weight and height."
            )

    return bmi


def render_feature(feat, artifact, sex, errors):

    stats = artifact["feature_stats"][feat]
    info = FEATURE_INFO.get(feat, {})

    if info.get("computed_from_weight_height"):
        return bmi_input(
            feat,
            stats,
            sex,
            errors
        )

    if stats["type"] == "numeric":
        return numeric_input(
            feat,
            stats,
            sex,
            errors
        )

    return categorical_input(
        feat,
        stats,
        sex,
        errors
    )


def get_feature_groups(artifact):

    groups = {
        "Anthropometrics": [],
        "Laboratory values": [],
        "Medical history": [],
        "Other": [],
    }

    for feat in artifact["feature_names"]:

        group = FEATURE_INFO.get(
            feat,
            {}
        ).get(
            "group",
            "Other"
        )

        groups.setdefault(group, [])
        groups[group].append(feat)

    return groups


def build_input_row(artifact, sex):

    values = {}
    errors = []

    groups = get_feature_groups(artifact)

    descriptions = {
        "Anthropometrics":
            "Basic demographic and anthropometric measurements.",

        "Laboratory values":
            "Routine laboratory parameters used by this model.",

        "Medical history":
            "Relevant questionnaire and medication information.",

        "Other":
            "Additional model variables.",
    }

    for group_name, features in groups.items():

        if not features:
            continue

        st.markdown(
            f'<div class="section-heading">{group_name}</div>',
            unsafe_allow_html=True,
        )

        st.markdown(
            f'<div class="section-description">'
            f'{descriptions[group_name]}'
            f'</div>',
            unsafe_allow_html=True,
        )

        # BMI receives full width because it contains
        # weight + height internally.
        regular_features = [
            f for f in features
            if not FEATURE_INFO.get(
                f, {}
            ).get(
                "computed_from_weight_height"
            )
        ]

        bmi_features = [
            f for f in features
            if FEATURE_INFO.get(
                f, {}
            ).get(
                "computed_from_weight_height"
            )
        ]

        # Regular features in two columns
        for i in range(
            0,
            len(regular_features),
            2
        ):

            cols = st.columns(2)

            pair = regular_features[
                i:i + 2
            ]

            for column, feat in zip(
                cols,
                pair
            ):
                with column:
                    values[feat] = (
                        render_feature(
                            feat,
                            artifact,
                            sex,
                            errors
                        )
                    )

        # BMI / calculated features
        for feat in bmi_features:
            values[feat] = render_feature(
                feat,
                artifact,
                sex,
                errors
            )

        st.markdown(
            '<div class="soft-divider"></div>',
            unsafe_allow_html=True,
        )

    row = pd.DataFrame(
        [values]
    )[artifact["feature_names"]]

    return row, errors


# ============================================================
# RESULT DISPLAY
# ============================================================

def render_result(
    probability,
    threshold
):

    probability_pct = probability * 100
    threshold_pct = threshold * 100

    above_threshold = (
        probability >= threshold
    )

    if above_threshold:
        classification = (
            "Pre-diabetes / Diabetes risk"
        )
        badge_class = "status-above"
        badge_text = (
            "Above decision threshold"
        )
    else:
        classification = "Normal"
        badge_class = "status-below"
        badge_text = (
            "Below decision threshold"
        )

    st.html(
        f"""
        <div class="result-card">

            <div class="result-label">
                Predicted probability of Pre-/DM
            </div>

            <div class="result-probability">
                {probability_pct:.1f}%
            </div>

            <div class="result-classification">
                Classification:
                <strong>{classification}</strong>
            </div>

            <div class="{badge_class}">
                {badge_text}
            </div>

            <div class="result-note">
                Decision threshold:
                {threshold_pct:.0f}%.
                This output represents model prediction
                only and should not be interpreted as a
                clinical diagnosis.
            </div>

        </div>
        """
    )

    st.progress(
        min(
            max(probability, 0.0),
            1.0
        )
    )


# ============================================================
# MAIN APP
# ============================================================

def render_single_patient_tab():

    # --------------------------------------------------------
    # Sex selection
    # --------------------------------------------------------

    st.html("""
    <div class="step-label">Step 1</div>
    <div class="section-title">Patient sex</div>
    <div class="section-subtitle">
        Select the appropriate sex-stratified model.
    </div>
    """)

    sex = st.segmented_control(
        "Patient sex",
        options=["Female", "Male"],
        default="Female",
        selection_mode="single",
        label_visibility="collapsed",
        width="stretch",
    )

    model_path = MODEL_PATHS[sex]

    if not os.path.exists(model_path):

        st.error(
            f"Model file not found: {model_path}\n\n"
            "Place both model files next to this script "
            "or set the MODEL_DIR environment variable."
        )

        return

    artifact = load_artifact(
        model_path
    )

    st.markdown(
        '<div class="soft-divider"></div>',
        unsafe_allow_html=True,
    )

    # --------------------------------------------------------
    # Main layout
    # --------------------------------------------------------

    left, right = st.columns(
        [2.2, 1],
        gap="large"
    )

    # ========================================================
    # LEFT: PATIENT DATA
    # ========================================================

    with left:

        st.markdown(
            '<div class="section-heading">'
            '2. Enter patient data'
            '</div>',
            unsafe_allow_html=True,
        )

        st.caption(
            f"This {sex.lower()} model uses "
            f"{len(artifact['feature_names'])} variables."
        )

        st.write("")

        X_row, input_errors = build_input_row(
            artifact,
            sex
        )

    # ========================================================
    # RIGHT: SETTINGS
    # ========================================================

    with right:

        with st.container(
            border=True
        ):

            st.markdown(
                "### Prediction settings"
            )

            threshold = st.slider(
                "Decision threshold",
                min_value=0.0,
                max_value=1.0,
                value=0.50,
                step=0.01,
                help=(
                    "Probability threshold used to "
                    "convert the model probability "
                    "into a binary classification."
                ),
            )

            st.caption(
                f"Current threshold: "
                f"{threshold * 100:.0f}%"
            )

        st.write("")

        with st.container(
            border=True
        ):

            st.markdown(
                "### Active model"
            )

            st.metric(
                "Sex stratum",
                artifact["sex_stratum"],
            )

            st.metric(
                "Features",
                len(
                    artifact[
                        "feature_names"
                    ]
                ),
            )

            st.metric(
                "Tuning CV AUC",
                f"{artifact['cv_auc_mean']:.3f}",
            )

            with st.expander(
                "Model details"
            ):

                st.write(
                    "**Features used**"
                )

                for feature in (
                    artifact[
                        "feature_names"
                    ]
                ):
                    st.write(
                        f"\u2022 {field_label(feature)}"
                    )

                st.caption(
                    "The displayed CV AUC is the "
                    "tuning-time 5-fold CV estimate "
                    "of the deployment model. "
                    "Use the isolated nested-CV "
                    "results for reporting model "
                    "generalization performance."
                )

    # --------------------------------------------------------
    # Prediction button
    # --------------------------------------------------------

    st.write("")

    st.markdown(
        '<div class="section-heading">'
        '3. Generate prediction'
        '</div>',
        unsafe_allow_html=True,
    )

    st.caption(
        "The entered values will be processed using "
        "the same preprocessing objects stored with the model."
    )

    if input_errors:
        st.error(
            "Fix the highlighted field(s) above before calculating:\n\n"
            + "\n".join(f"- {e}" for e in input_errors)
        )

    predict_clicked = st.button(
        "Calculate glycemic risk",
        type="primary",
        use_container_width=True,
        disabled=bool(input_errors),
    )

    # --------------------------------------------------------
    # Result
    # --------------------------------------------------------

    if predict_clicked and not input_errors:

        probability = predict(
            artifact,
            X_row
        )

        st.write("")

        st.markdown(
            '<div class="section-heading">'
            'Prediction result'
            '</div>',
            unsafe_allow_html=True,
        )

        render_result(
            probability,
            threshold
        )

        with st.expander(
            "View input values used for prediction"
        ):

            input_display = (
                X_row
                .T
                .reset_index()
            )

            input_display.columns = [
                "Feature",
                "Value"
            ]

            input_display[
                "Feature"
            ] = input_display[
                "Feature"
            ].map(
                lambda x:
                FEATURE_INFO.get(
                    x,
                    {}
                ).get(
                    "label",
                    x
                )
            )

            st.dataframe(
                input_display,
                hide_index=True,
                use_container_width=True,
            )


def main():

    # --------------------------------------------------------
    # Header
    # --------------------------------------------------------

    st.html("""
    <div class="app-header">

        <div class="header-top">
            <div>
                <div class="app-title">
                    Glycemic Risk Screening
                </div>

                <div class="app-subtitle">
                    Sex-specific machine-learning screening for
                    prediabetes and diabetes risk in trauma patients
                </div>
            </div>

            <div class="research-badge">
                RESEARCH USE ONLY
            </div>
        </div>

        <div class="disclaimer">
            <span class="disclaimer-dot"></span>
            Internal validation tool \u00b7 Not a clinical diagnostic device \u00b7
            Model predictions do not replace HbA1c testing or clinical assessment
        </div>

    </div>
    """)

    tab_single, tab_batch = st.tabs(
        ["\U0001f9cd Single patient", "\U0001f4cb Batch upload (Excel)"]
    )

    with tab_single:
        render_single_patient_tab()

    with tab_batch:
        render_batch_tab()


if __name__ == "__main__":
    main()